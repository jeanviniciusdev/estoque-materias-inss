import re
from io import BytesIO
import openpyxl
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.views import LoginView
from django.http import JsonResponse, HttpResponse
from django.db import models as dj_models
from .models import Material, Movimento
from .forms import MaterialForm, MovimentoForm
from .serializers import MaterialSerializer
from rest_framework.decorators import api_view
from django.utils.dateparse import parse_date
import csv
from django.contrib import messages
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST


class CustomLoginView(LoginView):
    template_name = 'registration/login.html'


@login_required
def dashboard(request):
    # top 5 materiais abaixo do mínimo
    alertas_top = Material.objects.filter(quantidade__lt=dj_models.F('minimo')).order_by('nome')[:5]
    # próximos 5 empréstimos ordenados por data_devolucao
    emprestimos_soon = Movimento.objects.filter(tipo='EMPRESTIMO').order_by('data_devolucao')[:5]
    return render(request, 'stock/dashboard.html', {
        'alertas_top': alertas_top,
        'emprestimos_soon': emprestimos_soon,
    })


@login_required
def alertas_completos(request):
    alertas = Material.objects.filter(quantidade__lt=dj_models.F('minimo')).order_by('nome')
    return render(request, 'stock/alertas_completos.html', {'alertas': alertas})


@login_required
def material_list(request):
    materiais = Material.objects.all()
    return render(request, 'stock/materials_list.html', {'materiais': materiais})


@login_required
def material_create(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES)
        if form.is_valid():
            mat = form.save()
            # Cria um movimento inicial se o material tiver quantidade
            if mat.quantidade and mat.quantidade > 0:
                mv = Movimento(material=mat, tipo='ADICAO', quantidade=mat.quantidade)
                if request.user.is_authenticated:
                    mv.usuario = request.user
                mv.save(adjust_material=False)
            return redirect('materials_list')
    else:
        form = MaterialForm()
    return render(request, 'stock/material_form.html', {'form': form})


@login_required
def material_edit(request, pk):
    mat = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        form = MaterialForm(request.POST, request.FILES, instance=mat)
        if request.POST.get('remove_image') == '1':
            if mat.imagem:
                mat.imagem.delete(save=False)
            mat.imagem = None
        if form.is_valid():
            form.save()
            return redirect('materials_list')
    else:
        form = MaterialForm(instance=mat)
    return render(request, 'stock/material_form.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_staff)
def material_delete(request, pk):
    mat = get_object_or_404(Material, pk=pk)
    if request.method == 'POST':
        motivo = (request.POST.get('motivo') or '').strip()
        if motivo:
            nota = f"nome={mat.nome}; motivo={motivo}"
        else:
            nota = f"nome={mat.nome}"
        mv = Movimento(
            material=mat,
            tipo='DELETE',
            quantidade=mat.quantidade or 0,
            nota=nota
        )
        if request.user.is_authenticated:
            mv.usuario = request.user
        mv.save(adjust_material=False)
        mat.delete()
        messages.success(request, f'Material "{mat.nome}" excluído.')
        return redirect('materials_list')
    return render(request, 'stock/material_confirm_delete.html', {'material': mat})


@login_required
def movimento_create(request):
    if request.method == 'POST':
        form = MovimentoForm(request.POST)
        if form.is_valid():
            mv = form.save(commit=False)
            if request.user.is_authenticated:
                mv.usuario = request.user
            # salvar usando default adjust_material=True para aplicar alteração de quantidade
            mv.save()
            messages.success(request, 'Movimento registrado.')
            return redirect('materials_list')
    else:
        form = MovimentoForm()
    return render(request, 'stock/movimento_form.html', {'form': form})

@api_view(['GET'])
def api_materials(request):
    materiais = Material.objects.all().order_by('nome')
    data = [
        {
            'id': m.id,
            'nome': m.nome,
            'quantidade': m.quantidade,
            'minimo': m.minimo,
            'imagem_url': (m.imagem.url if getattr(m, 'imagem', None) else None)
        }
        for m in materiais
    ]
    return JsonResponse(data, safe=False)

@login_required
def emprestimos_list(request):
    emprestimos = Movimento.objects.filter(
        tipo__in=['EMPRESTIMO', 'DEVOLVIDO']
    ).select_related('material', 'usuario').order_by('data_devolucao')

    # adiciona atributos temporários para exibição (não altera o modelo)
    for e in emprestimos:
        if e.tipo == 'DEVOLVIDO':
            e.status_label = '✅ Concluído'
            e.status_color = 'green'
        elif e.data_devolucao and e.data_devolucao < timezone.now().date():
            e.status_label = '⚠️ Atrasado'
            e.status_color = 'red'
        else:
            e.status_label = '⏳ Em andamento'
            e.status_color = 'orange'

    return render(request, 'stock/emprestimos_list.html', {'emprestimos': emprestimos})

@login_required
@user_passes_test(lambda u: u.is_staff)
def emprestimo_delete(request, pk):
    emprestimo = get_object_or_404(Movimento, pk=pk, tipo='EMPRESTIMO')
    if request.method == 'POST':
        emprestimo.delete()
        messages.success(request, f'Empréstimo de "{emprestimo.material.nome if emprestimo.material else "item"}" removido com sucesso.')
        return redirect('emprestimos_list')
    return render(request, 'stock/emprestimo_confirm_delete.html', {'emprestimo': emprestimo})



@login_required
def export_emprestimos(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    qs = Movimento.objects.filter(
        tipo__in=['EMPRESTIMO', 'DEVOLVIDO']
    ).select_related('material', 'usuario').order_by('data_devolucao')

    # adiciona status calculado para exportação
    for e in qs:
        if e.tipo == 'DEVOLVIDO':
            e.status_label = '✅ Concluído'
        elif e.data_devolucao and e.data_devolucao < timezone.now().date():
            e.status_label = '⚠️ Atrasado'
        else:
            e.status_label = '⏳ Em andamento'

    # --- XLSX ---
    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empréstimos"
        headers = [
            'ID', 'Data movimento', 'Usuário', 'Material',
            'Quantidade', 'Data prevista devolução', 'Nota', 'Status'
        ]
        ws.append(headers)

        for mv in qs:
            mat_nome = mv.material.nome if mv.material else (mv.nota or '')
            user = mv.usuario.username if mv.usuario else ''
            ws.append([
                mv.id,
                mv.criado.strftime('%Y-%m-%d'),
                user,
                mat_nome,
                mv.quantidade,
                mv.data_devolucao.strftime('%Y-%m-%d') if mv.data_devolucao else '',
                mv.nota,
                mv.status_label
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        resp = HttpResponse(
            stream.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        resp['Content-Disposition'] = 'attachment; filename="emprestimos.xlsx"'
        return resp

    # --- CSV ---
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="emprestimos.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, lineterminator='\r\n')
    writer.writerow([
        'ID', 'Data movimento', 'Usuário', 'Material',
        'Quantidade', 'Data prevista devolução', 'Nota', 'Status'
    ])

    for mv in qs:
        mat_nome = mv.material.nome if mv.material else (mv.nota or '')
        user = mv.usuario.username if mv.usuario else ''
        writer.writerow([
            mv.id,
            mv.criado.strftime('%Y-%m-%d'),
            user,
            mat_nome,
            mv.quantidade,
            mv.data_devolucao.strftime('%Y-%m-%d') if mv.data_devolucao else '',
            mv.nota,
            mv.status_label
        ])
    return response

@login_required
def export_materials_csv(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    materiais = Material.objects.all().order_by('nome')

    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Materiais"
        headers = ['ID', 'Nome', 'Descrição', 'Quantidade', 'Mínimo']
        ws.append(headers)
        for m in materiais:
            ws.append([m.id, m.nome, m.descricao, m.quantidade, m.minimo])
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        response = HttpResponse(stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="materiais.xlsx"'
        return response

    # fallback CSV (comportamento existente)
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="materiais.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, lineterminator='\r\n')
    writer.writerow(['ID', 'Nome', 'Descrição', 'Quantidade', 'Quantidade ideal'])
    for m in materiais:
        writer.writerow([m.id, m.nome, m.descricao, m.quantidade, m.minimo])
    return response


@login_required
def export_movements_csv(request):
    qs = Movimento.objects.select_related('material').all().order_by('-criado')

    start = request.GET.get('start')
    end = request.GET.get('end')

    if start:
        try:
            sd = parse_date(start)
            if sd:
                qs = qs.filter(criado__date__gte=sd)
        except Exception:
            pass

    if end:
        try:
            ed = parse_date(end)
            if ed:
                qs = qs.filter(criado__date__lte=ed)
        except Exception:
            pass

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="movimentos.csv"'
    response.write('\ufeff')

    writer = csv.writer(response, lineterminator='\r\n')
    # material_id removido do cabeçalho
    writer.writerow(['id', 'data', 'usuario', 'material_nome', 'tipo', 'quantidade', 'nota'])

    for mv in qs:
        usuario = mv.usuario.username if getattr(mv, 'usuario', None) else ''
        data_str = mv.criado.strftime('%d/%m/%Y') if getattr(mv, 'criado', None) else ''

        if mv.material is not None:
            material_nome = mv.material.nome
            nota_csv = mv.nota or ''
        else:
            material_nome = ''
            nota_csv = ''
            if mv.nota:
                # tenta extrair motivo da nota estruturada "nome=...; motivo=..."
                m_motivo = re.search(r'motivo=([^;]+)', mv.nota)
                if m_motivo:
                    nota_csv = m_motivo.group(1).strip()
                else:
                    # se não encontrar motivo, usa a nota inteira
                    nota_csv = mv.nota

        writer.writerow([
            mv.id,
            data_str,
            usuario,
            material_nome,
            mv.tipo,
            mv.quantidade,
            nota_csv
        ])

    return response


@login_required
def export_alertas_csv(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    alertas = Material.objects.filter(quantidade__lt=dj_models.F('minimo')).order_by('nome')

    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Alertas"
        headers = ['ID', 'Nome', 'Descrição', 'Quantidade Atual', 'Mínimo Necessário']
        ws.append(headers)
        for mat in alertas:
            ws.append([mat.id, mat.nome, mat.descricao, mat.quantidade, mat.minimo])
        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        response = HttpResponse(stream.read(),
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="materiais_alerta.xlsx"'
        return response

    # CSV fallback
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="materiais_alerta.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, lineterminator='\r\n')
    writer.writerow(['ID', 'Nome', 'Descrição', 'Quantidade Atual', 'Quantidade ideal'])
    
    for mat in alertas:
        writer.writerow([mat.id, mat.nome, mat.descricao, mat.quantidade, mat.minimo])
    return response

@login_required
def emprestimos_list(request):
    emprestimos = Movimento.objects.filter(tipo__in=['EMPRESTIMO', 'DEVOLVIDO']).select_related('material', 'usuario').order_by('data_devolucao')

    for e in emprestimos:
        if e.tipo == 'DEVOLVIDO':
            e.status_display = '✅ Concluído'
            e.status_color = 'green'
        elif e.data_devolucao and e.data_devolucao < timezone.now().date():
            e.status_display = '⚠️ Atrasado'
            e.status_color = 'red'
        else:
            e.status_display = '⏳ Em andamento'
            e.status_color = 'orange'

    return render(request, 'stock/emprestimos_list.html', {'emprestimos': emprestimos})


@login_required
@require_POST
def concluir_emprestimo(request, id):
    """
    Marca um empréstimo como concluído (finalizado). Se já estiver concluído,
    avisa e redireciona sem levantar 404.
    """
    emprestimo = get_object_or_404(Movimento, id=id)
    if emprestimo.tipo != 'DEVOLVIDO':
        emprestimo.tipo = 'DEVOLVIDO'
        # opcional: registrar a data de devolução como hoje
        emprestimo.data_devolucao = timezone.now().date()
        emprestimo.save()
        messages.success(request, f'O empréstimo do material "{emprestimo.material.nome if emprestimo.material else "item"}" foi concluído.')
    else:
        messages.info(request, f'O empréstimo do material "{emprestimo.material.nome if emprestimo.material else "item"}" já está concluído.')
    return redirect('emprestimos_list')



@login_required
@require_POST
def deletar_emprestimo(request, id):
    """
    Exclui um empréstimo da base de dados.
    """
    emprestimo = get_object_or_404(Movimento, id=id, tipo__in=['EMPRESTIMO', 'DEVOLVIDO'])
    nome = emprestimo.material.nome if emprestimo.material else "—"
    emprestimo.delete()
    messages.success(request, f'Empréstimo de "{nome}" foi excluído.')
    return redirect('emprestimos_list')

@login_required
def export_emprestimos(request):
    fmt = (request.GET.get('format') or 'csv').lower()
    qs = Movimento.objects.filter(tipo__in=['EMPRESTIMO', 'DEVOLVIDO']).select_related('material', 'usuario').order_by('data_devolucao')

    # garante que cada movimento tenha status calculado
    for e in qs:
        if e.tipo == 'DEVOLVIDO':
            e.status_display = '✅ Concluído'
        elif e.data_devolucao and e.data_devolucao < timezone.now().date():
            e.status_display = '⚠️ Atrasado'
        else:
            e.status_display = '⏳ Em andamento'

    # --- XLSX ---
    if fmt == 'xlsx':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Empréstimos"
        headers = ['ID','Data movimento','Usuário','Material','Quantidade','Data prevista devolução','Nota','Status']
        ws.append(headers)

        for mv in qs:
            mat_nome = mv.material.nome if mv.material else (mv.nota or '')
            user = mv.usuario.username if mv.usuario else ''
            ws.append([
                mv.id,
                mv.criado.strftime('%Y-%m-%d'),
                user,
                mat_nome,
                mv.quantidade,
                mv.data_devolucao.strftime('%Y-%m-%d') if mv.data_devolucao else '',
                mv.nota,
                mv.status_display
            ])

        stream = BytesIO()
        wb.save(stream)
        stream.seek(0)
        resp = HttpResponse(stream.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        resp['Content-Disposition'] = 'attachment; filename="emprestimos.xlsx"'
        return resp

    # --- CSV ---
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="emprestimos.csv"'
    response.write('\ufeff')
    writer = csv.writer(response, lineterminator='\r\n')
    writer.writerow(['ID','Data movimento','Usuário','Material','Quantidade','Data prevista devolução','Nota','Status'])

    for mv in qs:
        mat_nome = mv.material.nome if mv.material else (mv.nota or '')
        user = mv.usuario.username if mv.usuario else ''
        writer.writerow([
            mv.id,
            mv.criado.strftime('%Y-%m-%d'),
            user,
            mat_nome,
            mv.quantidade,
            mv.data_devolucao.strftime('%Y-%m-%d') if mv.data_devolucao else '',
            mv.nota,
            mv.status_display
        ])
    return response
